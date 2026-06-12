import base64
import io
import logging
import math
from typing import Any, Dict, List, Optional
from datetime import datetime, timedelta

import numpy as np
from sqlalchemy import select, desc, func, text
from sqlalchemy.ext.asyncio import AsyncSession

from db.postgres import AsyncSessionLocal
from db.postgres_models import (
    Company,
    StockQuote,
    Fundamental,
    NewsArticle,
    SocialSentiment,
    OptionChainRecord,
    Deal,
    MFHolding,
    Shareholding,
)

logger = logging.getLogger(__name__)

SYSTEM_PROMPT = """You are FinGraph AI, an expert financial analyst assistant for the Indian stock market (NSE/BSE).
You have access to real-time market data, fundamentals, risk analytics, knowledge graphs, predictions, and trading signals.

Guidelines:
- Always provide data-backed insights when possible
- Use tools to fetch real data before answering
- Be concise but thorough
- Include specific numbers (prices, percentages, ratios)
- When comparing stocks, use tables for clarity
- Mention data sources and timeframes
- For risk-related queries, highlight both upside and downside
- Never fabricate data — if a tool returns no data, say so
- Use Indian number format (lakhs, crores) where appropriate
"""


def _tool_schema(
    name: str,
    description: str,
    properties: Dict[str, Any],
    required: List[str],
) -> Dict[str, Any]:
    return {
        "type": "function",
        "function": {
            "name": name,
            "description": description,
            "parameters": {
                "type": "object",
                "properties": properties,
                "required": required,
            },
        },
    }


TOOL_DEFINITIONS = [
    _tool_schema(
        name="get_quote",
        description="Get the latest stock quote for an NSE/BSE symbol. Returns price, OHLCV, market cap.",
        properties={
            "symbol": {
                "type": "string",
                "description": "Stock symbol e.g. RELIANCE, TCS, INFY",
            }
        },
        required=["symbol"],
    ),
    _tool_schema(
        name="get_fundamentals",
        description="Get fundamental ratios for a stock: PE, PB, ROE, debt/equity, EPS, margins.",
        properties={
            "symbol": {
                "type": "string",
                "description": "Stock symbol",
            }
        },
        required=["symbol"],
    ),
    _tool_schema(
        name="get_historical",
        description="Get historical OHLCV data for a stock over a given period.",
        properties={
            "symbol": {"type": "string", "description": "Stock symbol"},
            "days": {
                "type": "integer",
                "description": "Number of days of history (default 30)",
                "default": 30,
            },
        },
        required=["symbol"],
    ),
    _tool_schema(
        name="get_option_chain",
        description="Get option chain data (calls and puts) for a stock with OI, volume, IV, Greeks.",
        properties={
            "symbol": {"type": "string", "description": "Stock symbol"},
        },
        required=["symbol"],
    ),
    _tool_schema(
        name="run_risk_analysis",
        description="Run risk analysis (VaR, Monte Carlo, stress test) for a stock or portfolio.",
        properties={
            "symbol": {"type": "string", "description": "Stock symbol"},
            "method": {
                "type": "string",
                "enum": ["var", "monte_carlo", "stress_test", "all"],
                "description": "Risk analysis method (default all)",
                "default": "all",
            },
        },
        required=["symbol"],
    ),
    _tool_schema(
        name="get_company_graph",
        description="Get knowledge graph relationships for a company: promoters, directors, sectors, subsidiaries.",
        properties={
            "symbol": {"type": "string", "description": "Stock symbol or company name"},
            "depth": {
                "type": "integer",
                "description": "Relationship depth (1 or 2, default 1)",
                "default": 1,
            },
        },
        required=["symbol"],
    ),
    _tool_schema(
        name="search_news",
        description="Search recent news articles. Optionally filter by stock symbol.",
        properties={
            "symbol": {
                "type": "string",
                "description": "Optional stock symbol to filter news",
            },
            "limit": {
                "type": "integer",
                "description": "Number of articles (default 10)",
                "default": 10,
            },
        },
        required=[],
    ),
    _tool_schema(
        name="run_screener",
        description="Screen stocks by fundamental criteria: sector, market cap range, PE, ROE, etc.",
        properties={
            "sector": {"type": "string", "description": "Sector filter"},
            "min_market_cap": {
                "type": "number",
                "description": "Minimum market cap in crores",
            },
            "max_pe": {"type": "number", "description": "Maximum PE ratio"},
            "min_roe": {"type": "number", "description": "Minimum ROE percentage"},
            "max_debt_equity": {
                "type": "number",
                "description": "Maximum debt-to-equity ratio",
            },
            "limit": {
                "type": "integer",
                "description": "Max results (default 20)",
                "default": 20,
            },
        },
        required=[],
    ),
    _tool_schema(
        name="get_prediction",
        description="Get ML prediction score and price target for a stock.",
        properties={
            "symbol": {"type": "string", "description": "Stock symbol"},
        },
        required=["symbol"],
    ),
    _tool_schema(
        name="get_trading_signals",
        description="Get technical trading signals (buy/sell/hold) for a stock.",
        properties={
            "symbol": {"type": "string", "description": "Stock symbol"},
        },
        required=["symbol"],
    ),
    _tool_schema(
        name="get_sentiment",
        description="Get social media sentiment analysis for a stock (Twitter, Reddit).",
        properties={
            "symbol": {"type": "string", "description": "Stock symbol"},
        },
        required=["symbol"],
    ),
    _tool_schema(
        name="compare_stocks",
        description="Compare fundamentals and performance of multiple stocks side by side.",
        properties={
            "symbols": {
                "type": "array",
                "items": {"type": "string"},
                "description": "List of stock symbols to compare",
            }
        },
        required=["symbols"],
    ),
    _tool_schema(
        name="screen_technicals",
        description="Screen stocks by technical indicators: RSI, MACD, Bollinger Bands. Returns stocks matching criteria.",
        properties={
            "indicator": {
                "type": "string",
                "enum": ["oversold_rsi", "overbought_rsi", "macd_bullish", "macd_bearish", "near_52w_high", "near_52w_low"],
                "description": "Technical criteria to screen for",
            },
            "sector": {
                "type": "string",
                "description": "Optional sector filter",
            },
            "limit": {
                "type": "integer",
                "description": "Max results (default 20)",
                "default": 20,
            },
        },
        required=["indicator"],
    ),
    _tool_schema(
        name="backtest_strategy",
        description="Backtest a simple trading strategy on historical data. Returns performance metrics.",
        properties={
            "symbol": {"type": "string", "description": "Stock symbol"},
            "strategy": {
                "type": "string",
                "enum": ["sma_crossover", "rsi_reversal", "buy_hold", "momentum"],
                "description": "Strategy to backtest",
            },
            "days": {
                "type": "integer",
                "description": "Historical period in days (default 365)",
                "default": 365,
            },
        },
        required=["symbol", "strategy"],
    ),
    _tool_schema(
        name="search_documents",
        description="Semantic search over news articles and research. Returns relevant documents matching the query.",
        properties={
            "query": {"type": "string", "description": "Search query in natural language"},
            "limit": {
                "type": "integer",
                "description": "Max results (default 10)",
                "default": 10,
            },
        },
        required=["query"],
    ),
    _tool_schema(
        name="get_institutional_activity",
        description="Get FII/DII activity, bulk/block deals, and mutual fund holdings changes.",
        properties={
            "symbol": {
                "type": "string",
                "description": "Optional stock symbol to filter",
            },
            "days": {
                "type": "integer",
                "description": "Lookback period in days (default 30)",
                "default": 30,
            },
        },
        required=[],
    ),
    _tool_schema(
        name="get_earnings_calendar",
        description="Get upcoming quarterly earnings dates and historical earnings surprises.",
        properties={
            "symbol": {
                "type": "string",
                "description": "Optional stock symbol",
            },
            "days_ahead": {
                "type": "integer",
                "description": "Days to look ahead (default 30)",
                "default": 30,
            },
        },
        required=[],
    ),
    # ── Excel Tools ──────────────────────────────────────────────
    _tool_schema(
        name="export_analysis_to_excel",
        description="Export financial analysis (comps, DCF, ratios) as an Excel file. Returns base64-encoded .xlsx.",
        properties={
            "symbol": {"type": "string", "description": "Stock symbol to analyze"},
            "analysis_type": {
                "type": "string",
                "enum": ["comps", "dcf_summary", "ratios", "historical_prices", "full"],
                "description": "What kind of analysis to export",
            },
        },
        required=["symbol", "analysis_type"],
    ),
    _tool_schema(
        name="import_financial_data",
        description="Import financial data from base64-encoded Excel or CSV content. Returns parsed records.",
        properties={
            "content": {"type": "string", "description": "Base64-encoded file content"},
            "file_type": {
                "type": "string",
                "enum": ["xlsx", "csv"],
                "description": "File format (default xlsx)",
                "default": "xlsx",
            },
        },
        required=["content"],
    ),
    # ── CFA-Level Tools ──────────────────────────────────────────
    _tool_schema(
        name="run_dcf_valuation",
        description="Full DCF valuation: projects free cash flow, calculates terminal value, discounts to present value using WACC, returns intrinsic value per share.",
        properties={
            "symbol": {"type": "string", "description": "Stock symbol"},
            "growth_rate": {
                "type": "number",
                "description": "Projected FCF growth rate % (default 10)",
                "default": 10,
            },
            "terminal_growth": {
                "type": "number",
                "description": "Terminal/perpetual growth rate % (default 3)",
                "default": 3,
            },
            "discount_rate": {
                "type": "number",
                "description": "WACC / discount rate % (default 12)",
                "default": 12,
            },
            "projection_years": {
                "type": "integer",
                "description": "Number of projection years (default 5)",
                "default": 5,
            },
        },
        required=["symbol"],
    ),
    _tool_schema(
        name="run_comps_analysis",
        description="Trading comps analysis: compares a stock's valuation multiples (PE, PB, EV/EBITDA, EV/Revenue) against sector peers.",
        properties={
            "symbol": {"type": "string", "description": "Stock symbol"},
            "sector": {
                "type": "string",
                "description": "Optional sector override. If omitted, uses the company's sector.",
            },
            "limit": {
                "type": "integer",
                "description": "Number of peers (default 10)",
                "default": 10,
            },
        },
        required=["symbol"],
    ),
    _tool_schema(
        name="run_dupont_analysis",
        description="Decompose ROE into Net Profit Margin × Asset Turnover × Equity Multiplier. Shows the drivers of return on equity.",
        properties={
            "symbol": {"type": "string", "description": "Stock symbol"},
        },
        required=["symbol"],
    ),
    _tool_schema(
        name="calculate_wacc",
        description="Calculate Weighted Average Cost of Capital (WACC) for a stock using CAPM for cost of equity and debt yields.",
        properties={
            "symbol": {"type": "string", "description": "Stock symbol"},
            "risk_free_rate": {
                "type": "number",
                "description": "Risk-free rate in % (default 7.0 = Indian 10Y G-sec)",
                "default": 7.0,
            },
            "market_risk_premium": {
                "type": "number",
                "description": "Market risk premium in % (default 6.0)",
                "default": 6.0,
            },
        },
        required=["symbol"],
    ),
    _tool_schema(
        name="run_portfolio_optimization",
        description="Run Markovitz mean-variance optimization on a portfolio of stocks. Returns efficient frontier, optimal weights, Sharpe ratio.",
        properties={
            "symbols": {
                "type": "array",
                "items": {"type": "string"},
                "description": "List of stock symbols in portfolio (2-10 stocks)",
            },
            "risk_free_rate": {
                "type": "number",
                "description": "Risk-free rate in % (default 7.0)",
                "default": 7.0,
            },
        },
        required=["symbols"],
    ),
    _tool_schema(
        name="calculate_bond_metrics",
        description="Calculate bond valuation metrics: yield to maturity, Macaulay duration, modified duration, convexity.",
        properties={
            "face_value": {
                "type": "number",
                "description": "Bond face value in rupees (default 1000)",
                "default": 1000,
            },
            "coupon_rate": {
                "type": "number",
                "description": "Annual coupon rate in %",
            },
            "years_to_maturity": {
                "type": "number",
                "description": "Years remaining until maturity",
            },
            "current_price": {
                "type": "number",
                "description": "Current market price in rupees",
            },
            "payment_frequency": {
                "type": "integer",
                "description": "Coupon payments per year (1=annual, 2=semi-annual, default 2)",
                "default": 2,
            },
        },
        required=["coupon_rate", "years_to_maturity", "current_price"],
    ),
    _tool_schema(
        name="run_financial_ratio_analysis",
        description="Comprehensive financial ratio analysis across 5 categories: profitability, liquidity, leverage, efficiency, valuation.",
        properties={
            "symbol": {"type": "string", "description": "Stock symbol"},
        },
        required=["symbol"],
    ),
    _tool_schema(
        name="generate_screener_config",
        description="Generate a screener configuration from a natural language description. Returns structured JSON with filters, columns, and sort settings.",
        properties={
            "description": {
                "type": "string",
                "description": "Natural language description of what stocks to screen for",
            }
        },
        required=["description"],
    ),
]


async def _resolve_company(symbol: str) -> Optional[Company]:
    async with AsyncSessionLocal() as session:
        sym_upper = symbol.upper().strip()
        result = await session.execute(
            select(Company).where(
                (Company.symbol == sym_upper)
                | (Company.nse_code == sym_upper)
                | (Company.bse_code == sym_upper)
            )
        )
        return result.scalar_one_or_none()


async def tool_get_quote(symbol: str) -> Dict[str, Any]:
    try:
        company = await _resolve_company(symbol)
        if not company:
            return {"error": f"Company not found: {symbol}"}

        async with AsyncSessionLocal() as session:
            q = await session.execute(
                select(StockQuote)
                .where(StockQuote.company_id == company.id)
                .order_by(desc(StockQuote.timestamp))
                .limit(1)
            )
            quote = q.scalar_one_or_none()

            if not quote:
                return {
                    "symbol": company.symbol,
                    "name": company.name,
                    "message": "No quote data available",
                }

            return {
                "symbol": company.symbol,
                "name": company.name,
                "sector": company.sector,
                "price": quote.close,
                "open": quote.open,
                "high": quote.high,
                "low": quote.low,
                "close": quote.close,
                "volume": quote.volume,
                "vwap": quote.vwap,
                "turnover": quote.turnover,
                "market_cap": company.market_cap,
                "timestamp": quote.timestamp.isoformat() if quote.timestamp else None,
            }
    except Exception as e:
        logger.error(f"get_quote error: {e}")
        return {"error": str(e)}


async def tool_get_fundamentals(symbol: str) -> Dict[str, Any]:
    try:
        company = await _resolve_company(symbol)
        if not company:
            return {"error": f"Company not found: {symbol}"}

        async with AsyncSessionLocal() as session:
            q = await session.execute(
                select(Fundamental)
                .where(Fundamental.company_id == company.id)
                .order_by(desc(Fundamental.id))
                .limit(1)
            )
            fund = q.scalar_one_or_none()

            if not fund:
                return {
                    "symbol": company.symbol,
                    "name": company.name,
                    "message": "No fundamental data available",
                }

            return {
                "symbol": company.symbol,
                "name": company.name,
                "pe": fund.pe,
                "pb": fund.pb,
                "roe": fund.roe,
                "roce": fund.roce,
                "eps": fund.eps,
                "debt_equity": fund.debt_equity,
                "current_ratio": fund.current_ratio,
                "gross_margin": fund.gross_margin,
                "net_margin": fund.net_margin,
                "dividend_yield": fund.dividend_yield,
                "revenue": fund.revenue,
                "profit": fund.profit,
                "quarter": fund.quarter,
                "fiscal_year": fund.fiscal_year,
            }
    except Exception as e:
        logger.error(f"get_fundamentals error: {e}")
        return {"error": str(e)}


async def tool_get_historical(symbol: str, days: int = 30) -> Dict[str, Any]:
    try:
        company = await _resolve_company(symbol)
        if not company:
            return {"error": f"Company not found: {symbol}"}

        async with AsyncSessionLocal() as session:
            cutoff = datetime.now() - timedelta(days=days)
            q = await session.execute(
                select(StockQuote)
                .where(
                    StockQuote.company_id == company.id,
                    StockQuote.timestamp >= cutoff,
                )
                .order_by(desc(StockQuote.timestamp))
                .limit(min(days, 365))
            )
            quotes = q.scalars().all()

            if not quotes:
                return {"symbol": company.symbol, "message": "No historical data"}

            return {
                "symbol": company.symbol,
                "name": company.name,
                "period_days": days,
                "data_points": len(quotes),
                "prices": [
                    {
                        "date": q.timestamp.isoformat() if q.timestamp else None,
                        "open": q.open,
                        "high": q.high,
                        "low": q.low,
                        "close": q.close,
                        "volume": q.volume,
                    }
                    for q in quotes
                ],
            }
    except Exception as e:
        logger.error(f"get_historical error: {e}")
        return {"error": str(e)}


async def tool_get_option_chain(symbol: str) -> Dict[str, Any]:
    try:
        async with AsyncSessionLocal() as session:
            q = await session.execute(
                select(OptionChainRecord)
                .where(OptionChainRecord.symbol == symbol.upper())
                .order_by(OptionChainRecord.strike_price, OptionChainRecord.option_type)
                .limit(100)
            )
            options = q.scalars().all()

            if not options:
                return {"symbol": symbol, "message": "No option chain data available"}

            calls = [o for o in options if o.option_type == "CE"]
            puts = [o for o in options if o.option_type == "PE"]

            return {
                "symbol": symbol.upper(),
                "calls": [
                    {
                        "strike": o.strike_price,
                        "last": o.last_price,
                        "oi": o.open_interest,
                        "volume": o.volume,
                        "iv": o.iv,
                    }
                    for o in calls
                ],
                "puts": [
                    {
                        "strike": o.strike_price,
                        "last": o.last_price,
                        "oi": o.open_interest,
                        "volume": o.volume,
                        "iv": o.iv,
                    }
                    for o in puts
                ],
            }
    except Exception as e:
        logger.error(f"get_option_chain error: {e}")
        return {"error": str(e)}


async def tool_run_risk_analysis(symbol: str, method: str = "all") -> Dict[str, Any]:
    try:
        from core.services.risk_engine import RiskEngine

        engine = RiskEngine()
        company = await _resolve_company(symbol)
        if not company:
            return {"error": f"Company not found: {symbol}"}

        results: Dict[str, Any] = {"symbol": company.symbol}

        if method in ("var", "all"):
            try:
                results["var"] = await engine.calculate_var(company.id)
            except Exception as e:
                results["var"] = {"error": str(e)}

        if method in ("monte_carlo", "all"):
            try:
                results["monte_carlo"] = await engine.monte_carlo_simulation(
                    company.id
                )
            except Exception as e:
                results["monte_carlo"] = {"error": str(e)}

        if method in ("stress_test", "all"):
            try:
                results["stress_test"] = await engine.stress_test(company.id)
            except Exception as e:
                results["stress_test"] = {"error": str(e)}

        return results
    except Exception as e:
        logger.error(f"run_risk_analysis error: {e}")
        return {"error": str(e)}


async def tool_get_company_graph(symbol: str, depth: int = 1) -> Dict[str, Any]:
    try:
        from core.services.graph_service import GraphService

        gs = GraphService()
        result = await gs.get_company_relationships(symbol, max_depth=depth)
        return result
    except Exception as e:
        logger.error(f"get_company_graph error: {e}")
        return {"error": str(e)}


async def tool_search_news(
    symbol: Optional[str] = None, limit: int = 10
) -> Dict[str, Any]:
    try:
        async with AsyncSessionLocal() as session:
            query = select(NewsArticle)

            if symbol:
                company = await _resolve_company(symbol)
                if company:
                    query = query.where(
                        NewsArticle.related_symbols.contains([company.symbol])
                    )

            query = query.order_by(desc(NewsArticle.published_at)).limit(limit)
            q = await session.execute(query)
            articles = q.scalars().all()

            return {
                "count": len(articles),
                "articles": [
                    {
                        "headline": a.headline,
                        "summary": a.summary,
                        "source": a.source,
                        "url": a.url,
                        "sentiment": a.sentiment,
                        "published_at": a.published_at.isoformat()
                        if a.published_at
                        else None,
                        "related_symbols": a.related_symbols,
                    }
                    for a in articles
                ],
            }
    except Exception as e:
        logger.error(f"search_news error: {e}")
        return {"error": str(e)}


async def tool_run_screener(
    sector: Optional[str] = None,
    min_market_cap: Optional[float] = None,
    max_pe: Optional[float] = None,
    min_roe: Optional[float] = None,
    max_debt_equity: Optional[float] = None,
    limit: int = 20,
) -> Dict[str, Any]:
    try:
        async with AsyncSessionLocal() as session:
            query = (
                select(Company, Fundamental)
                .join(Fundamental, Company.id == Fundamental.company_id)
                .where(Company.market_cap.isnot(None))
            )

            if sector:
                query = query.where(Company.sector.ilike(f"%{sector}%"))
            if min_market_cap:
                query = query.where(Company.market_cap >= min_market_cap * 1e7)
            if max_pe:
                query = query.where(
                    Fundamental.pe.isnot(None), Fundamental.pe <= max_pe
                )
            if min_roe:
                query = query.where(
                    Fundamental.roe.isnot(None), Fundamental.roe >= min_roe
                )
            if max_debt_equity:
                query = query.where(
                    Fundamental.debt_equity.isnot(None),
                    Fundamental.debt_equity <= max_debt_equity,
                )

            query = query.order_by(desc(Company.market_cap)).limit(limit)
            q = await session.execute(query)
            rows = q.all()

            return {
                "count": len(rows),
                "results": [
                    {
                        "symbol": company.symbol,
                        "name": company.name,
                        "sector": company.sector,
                        "market_cap_cr": round(
                            (company.market_cap or 0) / 1e7, 2
                        ),
                        "pe": fund.pe,
                        "pb": fund.pb,
                        "roe": fund.roe,
                        "debt_equity": fund.debt_equity,
                        "eps": fund.eps,
                    }
                    for company, fund in rows
                ],
            }
    except Exception as e:
        logger.error(f"run_screener error: {e}")
        return {"error": str(e)}


async def tool_get_prediction(symbol: str) -> Dict[str, Any]:
    try:
        from core.services.prediction import PredictionModel

        model = PredictionModel()
        result = await model.predict(symbol)
        return result
    except Exception as e:
        logger.error(f"get_prediction error: {e}")
        return {"error": str(e)}


async def tool_get_trading_signals(symbol: str) -> Dict[str, Any]:
    try:
        from core.services.signals import SignalGenerator

        gen = SignalGenerator()
        result = await gen.generate_signals(symbol)
        return result
    except Exception as e:
        logger.error(f"get_trading_signals error: {e}")
        return {"error": str(e)}


async def tool_get_sentiment(symbol: str) -> Dict[str, Any]:
    try:
        async with AsyncSessionLocal() as session:
            q = await session.execute(
                select(SocialSentiment)
                .where(SocialSentiment.symbol == symbol.upper())
                .order_by(desc(SocialSentiment.date))
                .limit(30)
            )
            sentiments = q.scalars().all()

            if not sentiments:
                return {
                    "symbol": symbol,
                    "message": "No sentiment data available",
                }

            avg_score = sum(s.sentiment_score or 0 for s in sentiments) / len(
                sentiments
            )
            total_mentions = sum(s.mention_count or 0 for s in sentiments)

            return {
                "symbol": symbol.upper(),
                "average_score": round(avg_score, 2),
                "total_mentions": total_mentions,
                "data_points": len(sentiments),
                "latest": {
                    "date": sentiments[0].date.isoformat()
                    if sentiments[0].date
                    else None,
                    "score": sentiments[0].sentiment_score,
                    "mentions": sentiments[0].mention_count,
                    "bullish": sentiments[0].bullish_count,
                    "bearish": sentiments[0].bearish_count,
                    "source": sentiments[0].source,
                },
                "trend": [
                    {
                        "date": s.date.isoformat() if s.date else None,
                        "score": s.sentiment_score,
                        "mentions": s.mention_count,
                    }
                    for s in sentiments[:10]
                ],
            }
    except Exception as e:
        logger.error(f"get_sentiment error: {e}")
        return {"error": str(e)}


async def tool_compare_stocks(symbols: List[str]) -> Dict[str, Any]:
    try:
        results = []
        for sym in symbols:
            company = await _resolve_company(sym)
            if not company:
                results.append({"symbol": sym, "error": "Not found"})
                continue

            async with AsyncSessionLocal() as session:
                fund_q = await session.execute(
                    select(Fundamental)
                    .where(Fundamental.company_id == company.id)
                    .order_by(desc(Fundamental.id))
                    .limit(1)
                )
                fund = fund_q.scalar_one_or_none()

                quote_q = await session.execute(
                    select(StockQuote)
                    .where(StockQuote.company_id == company.id)
                    .order_by(desc(StockQuote.timestamp))
                    .limit(1)
                )
                quote = quote_q.scalar_one_or_none()

                results.append(
                    {
                        "symbol": company.symbol,
                        "name": company.name,
                        "sector": company.sector,
                        "market_cap_cr": round(
                            (company.market_cap or 0) / 1e7, 2
                        ),
                        "price": quote.close if quote else None,
                        "pe": fund.pe if fund else None,
                        "pb": fund.pb if fund else None,
                        "roe": fund.roe if fund else None,
                        "eps": fund.eps if fund else None,
                        "debt_equity": fund.debt_equity if fund else None,
                        "net_margin": fund.net_margin if fund else None,
                    }
                )

        return {"comparison": results}
    except Exception as e:
        logger.error(f"compare_stocks error: {e}")
        return {"error": str(e)}


async def tool_screen_technicals(
    indicator: str,
    sector: Optional[str] = None,
    limit: int = 20,
) -> Dict[str, Any]:
    try:
        async with AsyncSessionLocal() as session:
            query = (
                select(Company, StockQuote)
                .join(StockQuote, Company.id == StockQuote.company_id)
                .where(Company.market_cap.isnot(None))
            )
            if sector:
                query = query.where(Company.sector.ilike(f"%{sector}%"))

            query = query.order_by(desc(StockQuote.timestamp))
            q = await session.execute(query)
            rows = q.all()

            company_quotes: Dict[str, List] = {}
            for company, quote in rows:
                if company.symbol not in company_quotes:
                    company_quotes[company.symbol] = []
                company_quotes[company.symbol].append(quote)

            results = []
            for symbol, quotes in company_quotes.items():
                if len(quotes) < 14:
                    continue
                prices = [q.close for q in reversed(quotes) if q.close]
                if len(prices) < 14:
                    continue

                company = next(r[0] for r in rows if r[0].symbol == symbol)

                if indicator == "oversold_rsi":
                    rsi = _calc_rsi(prices)
                    if rsi is not None and rsi < 30:
                        results.append({"symbol": symbol, "name": company.name, "rsi": round(rsi, 2), "price": prices[-1]})
                elif indicator == "overbought_rsi":
                    rsi = _calc_rsi(prices)
                    if rsi is not None and rsi > 70:
                        results.append({"symbol": symbol, "name": company.name, "rsi": round(rsi, 2), "price": prices[-1]})
                elif indicator == "macd_bullish":
                    macd_line, signal_line = _calc_macd(prices)
                    if macd_line is not None and signal_line is not None and macd_line > signal_line:
                        results.append({"symbol": symbol, "name": company.name, "macd": round(macd_line, 2), "signal": round(signal_line, 2), "price": prices[-1]})
                elif indicator == "macd_bearish":
                    macd_line, signal_line = _calc_macd(prices)
                    if macd_line is not None and signal_line is not None and macd_line < signal_line:
                        results.append({"symbol": symbol, "name": company.name, "macd": round(macd_line, 2), "signal": round(signal_line, 2), "price": prices[-1]})
                elif indicator == "near_52w_high":
                    high_52w = max(prices) if prices else 0
                    if high_52w > 0 and prices[-1] >= high_52w * 0.95:
                        results.append({"symbol": symbol, "name": company.name, "price": prices[-1], "high_52w": round(high_52w, 2), "pct_from_high": round((prices[-1] / high_52w - 1) * 100, 2)})
                elif indicator == "near_52w_low":
                    low_52w = min(prices) if prices else 0
                    if low_52w > 0 and prices[-1] <= low_52w * 1.05:
                        results.append({"symbol": symbol, "name": company.name, "price": prices[-1], "low_52w": round(low_52w, 2), "pct_from_low": round((prices[-1] / low_52w - 1) * 100, 2)})

                if len(results) >= limit:
                    break

            return {"indicator": indicator, "count": len(results), "results": results}
    except Exception as e:
        logger.error(f"screen_technicals error: {e}")
        return {"error": str(e)}


def _calc_rsi(prices: List[float], period: int = 14) -> Optional[float]:
    if len(prices) < period + 1:
        return None
    deltas = [prices[i] - prices[i - 1] for i in range(1, len(prices))]
    gains = [d if d > 0 else 0 for d in deltas[-period:]]
    losses = [-d if d < 0 else 0 for d in deltas[-period:]]
    avg_gain = sum(gains) / period
    avg_loss = sum(losses) / period
    if avg_loss == 0:
        return 100.0
    rs = avg_gain / avg_loss
    return 100 - (100 / (1 + rs))


def _calc_macd(prices: List[float]) -> tuple:
    if len(prices) < 26:
        return (None, None)
    ema12 = prices[0]
    ema26 = prices[0]
    for p in prices[1:]:
        ema12 = (p - ema12) * (2 / 13) + ema12
        ema26 = (p - ema26) * (2 / 27) + ema26
    macd_line = ema12 - ema26
    return (macd_line, macd_line * 0.8)


async def tool_backtest_strategy(
    symbol: str,
    strategy: str,
    days: int = 365,
) -> Dict[str, Any]:
    try:
        company = await _resolve_company(symbol)
        if not company:
            return {"error": f"Company not found: {symbol}"}

        async with AsyncSessionLocal() as session:
            cutoff = datetime.now() - timedelta(days=days)
            q = await session.execute(
                select(StockQuote)
                .where(StockQuote.company_id == company.id, StockQuote.timestamp >= cutoff)
                .order_by(StockQuote.timestamp)
            )
            quotes = q.scalars().all()

            if len(quotes) < 30:
                return {"symbol": symbol, "error": "Not enough historical data for backtest"}

            prices = [q.close for q in quotes if q.close]
            if len(prices) < 30:
                return {"symbol": symbol, "error": "Insufficient price data"}

            if strategy == "buy_hold":
                total_return = (prices[-1] - prices[0]) / prices[0] * 100
                max_dd = _max_drawdown(prices)
                return {
                    "symbol": symbol,
                    "strategy": "Buy and Hold",
                    "period_days": len(prices),
                    "total_return_pct": round(total_return, 2),
                    "max_drawdown_pct": round(max_dd, 2),
                    "start_price": round(prices[0], 2),
                    "end_price": round(prices[-1], 2),
                    "sharpe_ratio": round(_sharpe_ratio(prices), 2),
                }

            signals = []
            if strategy == "sma_crossover":
                sma_short = _sma(prices, 20)
                sma_long = _sma(prices, 50)
                if sma_short and sma_long:
                    signals = [1 if s > l else -1 for s, l in zip(sma_short, sma_long)]
            elif strategy == "momentum":
                signals = [1 if prices[i] > prices[i - 20] else -1 for i in range(20, len(prices))]
            elif strategy == "rsi_reversal":
                rsi_values = [_calc_rsi(prices[:i + 1]) for i in range(14, len(prices))]
                signals = [-1 if r and r > 70 else (1 if r and r < 30 else 0) for r in rsi_values]

            if not signals:
                return {"symbol": symbol, "strategy": strategy, "error": "Could not generate signals"}

            portfolio = 100000.0
            position = 0
            peak = portfolio
            max_dd = 0

            for i, sig in enumerate(signals):
                idx = i + (len(prices) - len(signals))
                if idx >= len(prices) or prices[idx] == 0:
                    continue
                daily_return = (prices[min(idx + 1, len(prices) - 1)] - prices[idx]) / prices[idx]
                if sig == 1:
                    portfolio *= (1 + daily_return)
                elif sig == -1:
                    portfolio *= (1 - daily_return)
                peak = max(peak, portfolio)
                dd = (peak - portfolio) / peak * 100
                max_dd = max(max_dd, dd)

            total_return = (portfolio - 100000) / 100000 * 100
            return {
                "symbol": symbol,
                "strategy": strategy,
                "period_days": len(prices),
                "total_return_pct": round(total_return, 2),
                "max_drawdown_pct": round(max_dd, 2),
                "final_portfolio": round(portfolio, 2),
                "sharpe_ratio": round(_sharpe_ratio(prices), 2),
                "trades": len([s for s in signals if s != 0]),
            }
    except Exception as e:
        logger.error(f"backtest_strategy error: {e}")
        return {"error": str(e)}


def _max_drawdown(prices: List[float]) -> float:
    peak = prices[0]
    max_dd = 0
    for p in prices:
        peak = max(peak, p)
        dd = (peak - p) / peak * 100
        max_dd = max(max_dd, dd)
    return max_dd


def _sharpe_ratio(prices: List[float], risk_free: float = 0.07) -> float:
    if len(prices) < 2:
        return 0
    returns = [(prices[i] - prices[i - 1]) / prices[i - 1] for i in range(1, len(prices))]
    avg_return = sum(returns) / len(returns)
    std = (sum((r - avg_return) ** 2 for r in returns) / len(returns)) ** 0.5
    if std == 0:
        return 0
    daily_rf = (1 + risk_free) ** (1 / 252) - 1
    return (avg_return - daily_rf) / std * (252 ** 0.5)


def _sma(data: List[float], period: int) -> List[float]:
    if len(data) < period:
        return []
    return [sum(data[i:i + period]) / period for i in range(len(data) - period + 1)]


async def tool_search_documents(query: str, limit: int = 10) -> Dict[str, Any]:
    try:
        keywords = query.lower().split()

        async with AsyncSessionLocal() as session:
            db_query = select(NewsArticle)
            conditions = []
            for kw in keywords[:5]:
                conditions.append(NewsArticle.headline.ilike(f"%{kw}%"))
            from sqlalchemy import or_
            if conditions:
                db_query = db_query.where(or_(*conditions))
            db_query = db_query.order_by(desc(NewsArticle.published_at)).limit(limit)
            q = await session.execute(db_query)
            articles = q.scalars().all()

            if not articles:
                db_query2 = select(NewsArticle).order_by(desc(NewsArticle.published_at)).limit(limit)
                q2 = await session.execute(db_query2)
                articles = q2.scalars().all()

            return {
                "query": query,
                "count": len(articles),
                "documents": [
                    {
                        "headline": a.headline,
                        "summary": a.summary,
                        "source": a.source,
                        "url": a.url,
                        "sentiment": a.sentiment,
                        "published_at": a.published_at.isoformat() if a.published_at else None,
                        "related_symbols": a.related_symbols,
                    }
                    for a in articles
                ],
            }
    except Exception as e:
        logger.error(f"search_documents error: {e}")
        return {"error": str(e)}


async def tool_get_institutional_activity(
    symbol: Optional[str] = None,
    days: int = 30,
) -> Dict[str, Any]:
    try:
        result_data = {"deals": [], "mf_changes": [], "shareholding": None}
        cutoff = datetime.now() - timedelta(days=days)

        async with AsyncSessionLocal() as session:
            if symbol:
                company = await _resolve_company(symbol)
                if not company:
                    return {"error": f"Company not found: {symbol}"}

                deal_q = await session.execute(
                    select(Deal)
                    .where(Deal.company_id == company.id, Deal.deal_date >= cutoff.date())
                    .order_by(desc(Deal.deal_date))
                    .limit(20)
                )
                deals = deal_q.scalars().all()
                result_data["deals"] = [
                    {
                        "type": d.deal_type,
                        "buyer": d.buyer_name,
                        "seller": d.seller_name,
                        "quantity": d.quantity,
                        "price": d.price,
                        "date": d.deal_date.isoformat() if d.deal_date else None,
                    }
                    for d in deals
                ]

                mf_q = await session.execute(
                    select(MFHolding)
                    .where(MFHolding.company_id == company.id)
                    .order_by(desc(MFHolding.year), desc(MFHolding.id))
                    .limit(10)
                )
                mf_holdings = mf_q.scalars().all()
                result_data["mf_changes"] = [
                    {
                        "mf_name": m.mf_name,
                        "quarter": m.quarter,
                        "year": m.year,
                        "quantity": m.quantity,
                        "change_qq": m.change_qq,
                    }
                    for m in mf_holdings
                ]

                sh_q = await session.execute(
                    select(Shareholding)
                    .where(Shareholding.company_id == company.id)
                    .order_by(desc(Shareholding.date))
                    .limit(1)
                )
                sh = sh_q.scalar_one_or_none()
                if sh:
                    result_data["shareholding"] = {
                        "promoter": sh.promoter,
                        "fii": sh.fii,
                        "dii": sh.dii,
                        "public": sh.public,
                        "date": sh.date.isoformat() if sh.date else None,
                    }

        return result_data
    except Exception as e:
        logger.error(f"get_institutional_activity error: {e}")
        return {"error": str(e)}


async def tool_get_earnings_calendar(
    symbol: Optional[str] = None,
    days_ahead: int = 30,
) -> Dict[str, Any]:
    try:
        result_data: Dict[str, Any] = {"upcoming": [], "historical": []}

        async with AsyncSessionLocal() as session:
            if symbol:
                company = await _resolve_company(symbol)
                if not company:
                    return {"error": f"Company not found: {symbol}"}

                fund_q = await session.execute(
                    select(Fundamental)
                    .where(Fundamental.company_id == company.id)
                    .order_by(desc(Fundamental.fiscal_year), desc(Fundamental.id))
                    .limit(8)
                )
                fundamentals = fund_q.scalars().all()

                result_data["historical"] = [
                    {
                        "symbol": company.symbol,
                        "name": company.name,
                        "quarter": f.quarter,
                        "fiscal_year": f.fiscal_year,
                        "revenue": f.revenue,
                        "profit": f.profit,
                        "eps": f.eps,
                    }
                    for f in fundamentals
                ]

                result_data["upcoming"] = [
                    {
                        "symbol": company.symbol,
                        "name": company.name,
                        "estimated_date": "Next quarter end (estimated)",
                        "note": "Exact date not available — based on quarterly cycle",
                    }
                ]
            else:
                recent_q = await session.execute(
                    select(Company, Fundamental)
                    .join(Fundamental, Company.id == Fundamental.company_id)
                    .order_by(desc(Fundamental.fiscal_year), desc(Fundamental.id))
                    .limit(20)
                )
                recent = recent_q.all()
                seen = set()
                for company, fund in recent:
                    if company.symbol not in seen:
                        seen.add(company.symbol)
                        result_data["historical"].append({
                            "symbol": company.symbol,
                            "name": company.name,
                            "latest_quarter": f"{fund.quarter} FY{fund.fiscal_year}",
                            "eps": fund.eps,
                            "revenue": fund.revenue,
                        })

        return result_data
    except Exception as e:
        logger.error(f"get_earnings_calendar error: {e}")
        return {"error": str(e)}


# ═══════════════════════════════════════════════════════════════
# Excel Tools
# ═══════════════════════════════════════════════════════════════


async def tool_export_analysis_to_excel(symbol: str, analysis_type: str) -> Dict[str, Any]:
    """Export financial analysis as base64 Excel file."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        company = await _resolve_company(symbol)
        if not company:
            return {"error": f"Company not found: {symbol}"}

        wb = openpyxl.Workbook()
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        def style_header(ws, cols: int):
            for col in range(1, cols + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border

        async with AsyncSessionLocal() as session:
            fund_q = await session.execute(
                select(Fundamental).where(Fundamental.company_id == company.id)
                .order_by(desc(Fundamental.id)).limit(1)
            )
            fund = fund_q.scalar_one_or_none()
            quotes_q = await session.execute(
                select(StockQuote).where(StockQuote.company_id == company.id)
                .order_by(desc(StockQuote.timestamp)).limit(365)
            )
            quotes = list(reversed(quotes_q.scalars().all()))

        if analysis_type in ("comps", "full"):
            ws = wb.active or wb.create_sheet("Comps")
            ws.title = "Comps"
            headers = ["Metric", "Value"]
            ws.append(headers)
            style_header(ws, 2)
            if fund:
                rows = [
                    ("PE Ratio", fund.pe), ("PB Ratio", fund.pb),
                    ("ROE %", fund.roe), ("ROCE %", fund.roce),
                    ("EPS", fund.eps), ("Debt/Equity", fund.debt_equity),
                    ("Dividend Yield %", fund.dividend_yield),
                    ("Net Margin %", fund.net_margin),
                    ("Revenue", fund.revenue), ("Profit", fund.profit),
                ]
                for label, val in rows:
                    ws.append([label, val])
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 20

        if analysis_type in ("historical_prices", "full") and quotes:
            ws2 = wb.create_sheet("Prices")
            headers = ["Date", "Open", "High", "Low", "Close", "Volume"]
            ws2.append(headers)
            style_header(ws2, 6)
            for q in quotes:
                dt = q.timestamp.strftime("%Y-%m-%d") if q.timestamp else ""
                ws2.append([dt, q.open, q.high, q.low, q.close, q.volume])
            ws2.column_dimensions['A'].width = 14
            for c in "BCDEF":
                ws2.column_dimensions[c].width = 14

        if analysis_type in ("dcf_summary", "full") and fund:
            ws3 = wb.create_sheet("DCF Summary")
            headers = ["Parameter", "Value"]
            ws3.append(headers)
            style_header(ws3, 2)
            fcf_est = (fund.profit or 0) * 0.7
            wacc = 12.0
            dcf_rows = [
                ("Estimated FCF (₹ Cr)", round(fcf_est / 1e7, 2) if fcf_est else "N/A"),
                ("WACC %", wacc),
                ("Projection Years", 5),
                ("Terminal Growth %", 3.0),
            ]
            for label, val in dcf_rows:
                ws3.append([label, val])
            ws3.column_dimensions['A'].width = 25
            ws3.column_dimensions['B'].width = 20

        if analysis_type in ("ratios", "full") and fund:
            ws4 = wb.create_sheet("Ratios")
            headers = ["Category", "Ratio", "Value"]
            ws4.append(headers)
            style_header(ws4, 3)
            ratio_rows = [
                ("Profitability", "Net Margin %", fund.net_margin),
                ("Profitability", "ROE %", fund.roe),
                ("Profitability", "ROCE %", fund.roce),
                ("Leverage", "Debt/Equity", fund.debt_equity),
                ("Valuation", "PE", fund.pe),
                ("Valuation", "PB", fund.pb),
                ("Valuation", "EPS", fund.eps),
                ("Valuation", "Dividend Yield %", fund.dividend_yield),
            ]
            for cat, name, val in ratio_rows:
                ws4.append([cat, name, val])
            for c in "ABC":
                ws4.column_dimensions[c].width = 20

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        b64 = base64.b64encode(buf.read()).decode()

        return {
            "symbol": symbol,
            "analysis_type": analysis_type,
            "file_name": f"{symbol}_{analysis_type}.xlsx",
            "file_size_bytes": len(b64),
            "content_base64": b64,
            "sheets": wb.sheetnames,
        }
    except ImportError:
        return {"error": "openpyxl not installed. Run: pip install openpyxl"}
    except Exception as e:
        logger.error(f"export_analysis_to_excel error: {e}")
        return {"error": str(e)}


async def tool_import_financial_data(content: str, file_type: str = "xlsx") -> Dict[str, Any]:
    """Parse base64-encoded Excel/CSV content into structured records."""
    try:
        raw = base64.b64decode(content)

        if file_type == "csv":
            import csv
            text = raw.decode("utf-8")
            reader = csv.DictReader(io.StringIO(text))
            records = [row for row in reader]
            return {"file_type": "csv", "row_count": len(records), "columns": list(records[0].keys()) if records else [], "records": records}

        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
        sheets = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [str(h) if h else f"col_{i}" for i, h in enumerate(rows[0])]
            data = []
            for row in rows[1:]:
                record = {}
                for i, val in enumerate(row):
                    if i < len(headers):
                        record[headers[i]] = val
                if any(v is not None for v in record.values()):
                    data.append(record)
            sheets[sheet_name] = {"headers": headers, "row_count": len(data), "data": data}
        return {"file_type": "xlsx", "sheets": sheets, "sheet_names": wb.sheetnames}
    except Exception as e:
        logger.error(f"import_financial_data error: {e}")
        return {"error": str(e)}


# ═══════════════════════════════════════════════════════════════
# CFA-Level Tools
# ═══════════════════════════════════════════════════════════════


async def tool_run_dcf_valuation(
    symbol: str,
    growth_rate: float = 10.0,
    terminal_growth: float = 3.0,
    discount_rate: float = 12.0,
    projection_years: int = 5,
) -> Dict[str, Any]:
    """Full DCF valuation — intrinsic value per share."""
    try:
        company = await _resolve_company(symbol)
        if not company:
            return {"error": f"Company not found: {symbol}"}

        async with AsyncSessionLocal() as session:
            fq = await session.execute(
                select(Fundamental).where(Fundamental.company_id == company.id)
                .order_by(desc(Fundamental.id)).limit(1)
            )
            fund = fq.scalar_one_or_none()
            pq = await session.execute(
                select(StockQuote).where(StockQuote.company_id == company.id)
                .order_by(desc(StockQuote.timestamp)).limit(1)
            )
            quote = pq.scalar_one_or_none()

        if not fund:
            return {"error": "No fundamental data available for DCF"}

        # Rough FCF estimate: profit * 0.7 (conversion ratio)
        base_fcf = (fund.profit or 0) * 0.7
        if base_fcf <= 0:
            return {"error": "Negative or zero profit — cannot run DCF"}

        gr = growth_rate / 100
        tg = terminal_growth / 100
        dr = discount_rate / 100

        projections = []
        cum_dcf = 0.0
        for year in range(1, projection_years + 1):
            fcf = base_fcf * ((1 + gr) ** year)
            discount_factor = 1 / ((1 + dr) ** year)
            pv = fcf * discount_factor
            cum_dcf += pv
            projections.append({
                "year": year,
                "projected_fcf_cr": round(fcf / 1e7, 2),
                "discount_factor": round(discount_factor, 4),
                "pv_of_fcf_cr": round(pv / 1e7, 2),
            })

        # Terminal value
        terminal_fcf = base_fcf * ((1 + gr) ** projection_years) * (1 + tg)
        terminal_value = terminal_fcf / (dr - tg) if dr > tg else 0
        pv_terminal = terminal_value / ((1 + dr) ** projection_years)

        total_ev = cum_dcf + pv_terminal

        # Subtract debt, add cash for equity value
        net_debt_estimate = (fund.debt_equity or 0.5) * (fund.revenue or base_fcf * 2) * 0.3
        equity_value = total_ev - net_debt_estimate

        # Shares outstanding estimate from market cap / price
        current_price = quote.close if quote else None
        market_cap = company.market_cap or (fund.pe or 20) * (fund.profit or base_fcf)
        shares_outstanding = market_cap / current_price if current_price and market_cap else 0

        intrinsic_value = equity_value / shares_outstanding if shares_outstanding > 0 else 0

        current_val = current_price or 0
        upside = ((intrinsic_value / current_val) - 1) * 100 if current_val > 0 else 0

        return {
            "symbol": symbol,
            "name": company.name,
            "valuation_date": datetime.now().strftime("%Y-%m-%d"),
            "assumptions": {
                "base_fcf_cr": round(base_fcf / 1e7, 2),
                "fcf_growth_rate_pct": growth_rate,
                "terminal_growth_pct": terminal_growth,
                "wacc_pct": discount_rate,
                "projection_years": projection_years,
            },
            "projections": projections,
            "terminal_value": {
                "terminal_fcf_cr": round(terminal_fcf / 1e7, 2),
                "terminal_value_cr": round(terminal_value / 1e7, 2),
                "pv_of_terminal_value_cr": round(pv_terminal / 1e7, 2),
            },
            "valuation": {
                "total_enterprise_value_cr": round(total_ev / 1e7, 2),
                "net_debt_estimate_cr": round(net_debt_estimate / 1e7, 2),
                "equity_value_cr": round(equity_value / 1e7, 2),
                "shares_outstanding_cr": round(shares_outstanding / 1e7, 2) if shares_outstanding else None,
                "intrinsic_value_per_share": round(intrinsic_value, 2),
                "current_price": current_val,
                "upside_pct": round(upside, 1),
                "verdict": "Undervalued" if upside > 15 else "Overvalued" if upside < -15 else "Fairly Valued",
            },
        }
    except Exception as e:
        logger.error(f"run_dcf_valuation error: {e}")
        return {"error": str(e)}


async def tool_run_comps_analysis(symbol: str, sector: Optional[str] = None, limit: int = 10) -> Dict[str, Any]:
    """Trading comps — compares a stock's multiples against sector peers."""
    try:
        company = await _resolve_company(symbol)
        if not company:
            return {"error": f"Company not found: {symbol}"}

        target_sector = sector or company.sector
        if not target_sector:
            return {"error": "No sector available for comps"}

        async with AsyncSessionLocal() as session:
            peers_q = await session.execute(
                select(Company, Fundamental)
                .join(Fundamental, Company.id == Fundamental.company_id)
                .where(Company.sector.ilike(f"%{target_sector}%"))
                .where(Company.market_cap.isnot(None))
                .order_by(desc(Company.market_cap))
                .limit(limit + 5)
            )
            rows = peers_q.all()

        peers = []
        for c, f in rows:
            if c.symbol == symbol:
                continue
            peers.append({
                "symbol": c.symbol, "name": c.name,
                "market_cap_cr": round((c.market_cap or 0) / 1e7, 2),
                "pe": f.pe, "pb": f.pb, "roe": f.roe,
                "eps": f.eps, "debt_equity": f.debt_equity,
            })

        peers = peers[:limit]

        def median(values):
            sv = sorted([v for v in values if v is not None])
            if not sv:
                return None
            return sv[len(sv) // 2]

        pe_vals = [p["pe"] for p in peers if p["pe"]]
        pb_vals = [p["pb"] for p in peers if p["pb"]]
        roe_vals = [p["roe"] for p in peers if p["roe"]]

        # Get the target company's own multiples
        async with AsyncSessionLocal() as session:
            fq = await session.execute(
                select(Fundamental).where(Fundamental.company_id == company.id)
                .order_by(desc(Fundamental.id)).limit(1)
            )
            fund = fq.scalar_one_or_none()

        target = {
            "symbol": company.symbol, "name": company.name,
            "market_cap_cr": round((company.market_cap or 0) / 1e7, 2),
            "pe": fund.pe if fund else None, "pb": fund.pb if fund else None,
            "roe": fund.roe if fund else None, "eps": fund.eps if fund else None,
            "debt_equity": fund.debt_equity if fund else None,
        }

        return {
            "target_sector": target_sector,
            "peer_count": len(peers),
            "target_company": target,
            "peer_median": {
                "pe": round(median(pe_vals), 2) if pe_vals else None,
                "pb": round(median(pb_vals), 2) if pb_vals else None,
                "roe": round(median(roe_vals), 2) if roe_vals else None,
            },
            "peer_min": {
                "pe": round(min(pe_vals), 2) if pe_vals else None,
                "pb": round(min(pb_vals), 2) if pb_vals else None,
            },
            "peer_max": {
                "pe": round(max(pe_vals), 2) if pe_vals else None,
                "pb": round(max(pb_vals), 2) if pb_vals else None,
            },
            "percentile_vs_peers": {
                "pe_cheaper_pct": round(sum(1 for p in pe_vals if p > (target["pe"] or 0)) / len(pe_vals) * 100, 1) if pe_vals and target.get("pe") else None,
                "roe_better_pct": round(sum(1 for p in roe_vals if p < (target["roe"] or 0)) / len(roe_vals) * 100, 1) if roe_vals and target.get("roe") else None,
            },
            "peers": peers,
        }
    except Exception as e:
        logger.error(f"run_comps_analysis error: {e}")
        return {"error": str(e)}


async def tool_run_dupont_analysis(symbol: str) -> Dict[str, Any]:
    """Decompose ROE into Net Profit Margin × Asset Turnover × Equity Multiplier."""
    try:
        company = await _resolve_company(symbol)
        if not company:
            return {"error": f"Company not found: {symbol}"}

        async with AsyncSessionLocal() as session:
            fq = await session.execute(
                select(Fundamental).where(Fundamental.company_id == company.id)
                .order_by(desc(Fundamental.id)).limit(1)
            )
            fund = fq.scalar_one_or_none()

        if not fund:
            return {"error": "No fundamental data available"}

        net_margin = (fund.net_margin or 0) / 100 if fund.net_margin else None
        roe = (fund.roe or 0) / 100 if fund.roe else None
        debt_equity = fund.debt_equity or 0

        # DuPont: ROE = NPM × AT × EM
        # If we know NPM and ROE, we can solve for AT × EM = ROE / NPM
        # EM = 1 + D/E
        equity_multiplier = 1 + debt_equity

        if net_margin and equity_multiplier > 0:
            asset_turnover = (roe / (net_margin * equity_multiplier)) if net_margin and equity_multiplier else None
        else:
            asset_turnover = None

        roe_pct = fund.roe
        npm_pct = fund.net_margin

        return {
            "symbol": symbol,
            "name": company.name,
            "formula": "ROE = Net Profit Margin × Asset Turnover × Equity Multiplier",
            "components": {
                "net_profit_margin_pct": npm_pct,
                "asset_turnover": round(asset_turnover, 3) if asset_turnover else None,
                "equity_multiplier": round(equity_multiplier, 2),
                "debt_to_equity": debt_equity,
            },
            "result": {
                "calculated_roe_pct": round(npm_pct * asset_turnover * equity_multiplier * 100, 2) if (npm_pct and asset_turnover) else None,
                "actual_roe_pct": roe_pct,
            },
            "interpretation": _dupont_interpretation(npm_pct, asset_turnover, equity_multiplier, roe_pct),
        }
    except Exception as e:
        logger.error(f"run_dupont_analysis error: {e}")
        return {"error": str(e)}


def _dupont_interpretation(npm, at, em, roe):
    parts = []
    if npm and npm > 15:
        parts.append("High profit margins suggest strong pricing power or cost advantage.")
    elif npm and npm < 5:
        parts.append("Thin margins indicate competitive pricing pressure or high costs.")
    if at and at > 1.0:
        parts.append("Efficient asset utilisation — generates significant revenue per rupee of assets.")
    elif at and at < 0.5:
        parts.append("Asset-heavy business model with lower turnover.")
    if em and em > 2.5:
        parts.append("Leverage-driven returns — high debt amplifies ROE but increases financial risk.")
    if em and em < 1.5:
        parts.append("Conservative capital structure — low debt reliance.")
    if roe and roe > 20:
        parts.append("Strong ROE indicates effective capital allocation.")
    return " ".join(parts) if parts else "Mixed signals — examine individual components."


async def tool_calculate_wacc(symbol: str, risk_free_rate: float = 7.0, market_risk_premium: float = 6.0) -> Dict[str, Any]:
    """Calculate WACC using CAPM for cost of equity."""
    try:
        company = await _resolve_company(symbol)
        if not company:
            return {"error": f"Company not found: {symbol}"}

        async with AsyncSessionLocal() as session:
            fq = await session.execute(
                select(Fundamental).where(Fundamental.company_id == company.id)
                .order_by(desc(Fundamental.id)).limit(1)
            )
            fund = fq.scalar_one_or_none()
            pq = await session.execute(
                select(StockQuote).where(StockQuote.company_id == company.id)
                .order_by(desc(StockQuote.timestamp)).limit(1)
            )
            quote = pq.scalar_one_or_none()

        if not fund:
            return {"error": "No fundamental data available"}

        # Estimate beta from sector
        sector_beta_map = {
            "Technology": 1.15, "Financial Services": 1.05, "Energy": 0.95,
            "Banking": 1.10, "IT": 1.15, "Pharma": 0.85, "FMCG": 0.65,
            "Auto": 1.20, "Capital Goods": 1.10, "Construction": 1.15,
            "Metals": 1.30, "Power": 0.75, "Telecom": 1.25, "Media": 1.10,
        }
        beta = sector_beta_map.get(company.sector or "", 1.0)

        rfr = risk_free_rate / 100
        mrp = market_risk_premium / 100

        # CAPM: Re = Rf + β × MRP
        cost_of_equity = rfr + beta * mrp

        # Cost of debt: risk-free rate + credit spread
        credit_spread = 0.02  # 200 bps for investment grade
        cost_of_debt = rfr + credit_spread

        # Tax rate estimate (Indian corporate)
        tax_rate = 0.25

        # Weights: D/(D+E) from debt/equity ratio
        debt_equity = fund.debt_equity or 0.5
        equity_weight = 1 / (1 + debt_equity)
        debt_weight = debt_equity / (1 + debt_equity)

        wacc = equity_weight * cost_of_equity + debt_weight * cost_of_debt * (1 - tax_rate)

        return {
            "symbol": symbol,
            "name": company.name,
            "sector": company.sector,
            "inputs": {
                "risk_free_rate_pct": risk_free_rate,
                "market_risk_premium_pct": market_risk_premium,
                "beta_estimated": round(beta, 2),
                "cost_of_equity_pct": round(cost_of_equity * 100, 2),
                "cost_of_debt_pretax_pct": round(cost_of_debt * 100, 2),
                "cost_of_debt_aftertax_pct": round(cost_of_debt * (1 - tax_rate) * 100, 2),
                "debt_to_equity": round(debt_equity, 2),
                "equity_weight_pct": round(equity_weight * 100, 1),
                "debt_weight_pct": round(debt_weight * 100, 1),
                "tax_rate_pct": tax_rate * 100,
            },
            "wacc_pct": round(wacc * 100, 2),
            "current_price": quote.close if quote else None,
            "interpretation": f"WACC of {round(wacc * 100, 1)}% implies the company must generate returns above this threshold to create value." if wacc else "",
        }
    except Exception as e:
        logger.error(f"calculate_wacc error: {e}")
        return {"error": str(e)}


async def tool_run_portfolio_optimization(symbols: List[str], risk_free_rate: float = 7.0) -> Dict[str, Any]:
    """Markowitz mean-variance optimization. Computes efficient frontier and optimal weights."""
    try:
        if len(symbols) < 2 or len(symbols) > 10:
            return {"error": "Portfolio must have 2-10 stocks"}

        price_data = {}
        for sym in symbols:
            company = await _resolve_company(sym)
            if not company:
                continue
            async with AsyncSessionLocal() as session:
                qq = await session.execute(
                    select(StockQuote).where(StockQuote.company_id == company.id)
                    .where(StockQuote.close.isnot(None))
                    .order_by(desc(StockQuote.timestamp)).limit(252)
                )
                quotes = list(reversed(qq.scalars().all()))
            if len(quotes) > 20:
                price_data[sym] = [q.close for q in quotes]

        valid_symbols = list(price_data.keys())
        if len(valid_symbols) < 2:
            return {"error": "Need at least 2 stocks with sufficient price history"}

        # Build returns matrix
        returns_matrix = []
        for sym in valid_symbols:
            prices = price_data[sym]
            ret = [(prices[i] - prices[i - 1]) / prices[i - 1] for i in range(1, len(prices))]
            returns_matrix.append(ret)

        min_len = min(len(r) for r in returns_matrix)
        returns_array = np.array([r[:min_len] for r in returns_matrix])

        # Mean returns and covariance
        mean_returns = np.mean(returns_array, axis=1)
        cov_matrix = np.cov(returns_array)

        rfr = risk_free_rate / 100

        num_assets = len(valid_symbols)
        num_portfolios = 5000
        results = []
        max_sharpe = -np.inf
        max_sharpe_weights = None

        for _ in range(num_portfolios):
            weights = np.random.random(num_assets)
            weights /= np.sum(weights)

            portfolio_return = np.sum(mean_returns * weights)
            portfolio_std = np.sqrt(np.dot(weights.T, np.dot(cov_matrix, weights)))
            sharpe = (portfolio_return - rfr) / portfolio_std if portfolio_std > 0 else 0

            results.append({
                "return_pct": round(portfolio_return * 100, 2),
                "risk_pct": round(portfolio_std * 100, 2),
                "sharpe": round(sharpe, 3),
            })

            if sharpe > max_sharpe:
                max_sharpe = sharpe
                max_sharpe_weights = weights

        # Find min variance
        min_var = min(results, key=lambda r: r["risk_pct"])

        optimal = {}
        if max_sharpe_weights is not None:
            optimal = {valid_symbols[i]: round(float(w) * 100, 1) for i, w in enumerate(max_sharpe_weights)}

        efficient_frontier = sorted(results, key=lambda r: r["risk_pct"])
        frontier_sample = efficient_frontier[::max(1, len(efficient_frontier) // 20)]

        return {
            "portfolio_size": len(valid_symbols),
            "symbols": valid_symbols,
            "risk_free_rate_pct": risk_free_rate,
            "optimal_portfolio": {
                "expected_return_pct": round(np.sum(mean_returns * max_sharpe_weights) * 100, 2) if max_sharpe_weights is not None else None,
                "expected_risk_pct": round(np.sqrt(np.dot(max_sharpe_weights.T, np.dot(cov_matrix, max_sharpe_weights))) * 100, 2) if max_sharpe_weights is not None else None,
                "sharpe_ratio": round(max_sharpe, 3),
                "weights_pct": optimal,
            },
            "min_variance_portfolio": {
                "risk_pct": min_var["risk_pct"],
                "return_pct": min_var["return_pct"],
            },
            "efficient_frontier": frontier_sample[:50],
            "interpretation": f"Optimal Sharpe ratio of {round(max_sharpe, 2)} suggests {'strong' if max_sharpe > 1 else 'moderate' if max_sharpe > 0.5 else 'weak'} risk-adjusted returns." if max_sharpe else "",
        }
    except Exception as e:
        logger.error(f"run_portfolio_optimization error: {e}")
        return {"error": str(e)}


async def tool_calculate_bond_metrics(
    face_value: float = 1000,
    coupon_rate: float = 0,
    years_to_maturity: float = 0,
    current_price: float = 0,
    payment_frequency: int = 2,
) -> Dict[str, Any]:
    """Calculate bond YTM, Macaulay duration, modified duration, convexity."""
    try:
        cr = coupon_rate / 100
        periods = int(years_to_maturity * payment_frequency)
        coupon = face_value * cr / payment_frequency
        ytm_guess = (coupon + (face_value - current_price) / periods) / ((face_value + current_price) / 2)

        # Newton-Raphson for YTM
        ytm = ytm_guess
        for _ in range(100):
            price_calc = 0
            dprice_dy = 0
            for t in range(1, periods + 1):
                price_calc += coupon / ((1 + ytm / payment_frequency) ** t)
                dprice_dy += -t * coupon / (payment_frequency * (1 + ytm / payment_frequency) ** (t + 1))
            price_calc += face_value / ((1 + ytm / payment_frequency) ** periods)
            dprice_dy += -periods * face_value / (payment_frequency * (1 + ytm / payment_frequency) ** (periods + 1))

            diff = price_calc - current_price
            if abs(diff) < 0.001:
                break
            if dprice_dy == 0:
                break
            ytm -= diff / dprice_dy

        ytm_annual = ytm * 100

        # Macaulay Duration
        mac_dur = 0
        price_check = 0
        for t in range(1, periods + 1):
            cf = coupon if t < periods else coupon + face_value
            pv = cf / ((1 + ytm / payment_frequency) ** t)
            price_check += pv
            mac_dur += t * pv
        if price_check > 0:
            mac_dur = mac_dur / price_check / payment_frequency  # in years

        # Modified Duration
        mod_dur = mac_dur / (1 + ytm / payment_frequency) if ytm > -payment_frequency else 0

        # Convexity
        convexity = 0
        for t in range(1, periods + 1):
            cf = coupon if t < periods else coupon + face_value
            pv = cf / ((1 + ytm / payment_frequency) ** t)
            convexity += t * (t + 1) * pv / ((1 + ytm / payment_frequency) ** 2)
        if price_check > 0:
            convexity = convexity / price_check / (payment_frequency ** 2)

        return {
            "face_value": face_value,
            "coupon_rate_pct": coupon_rate,
            "years_to_maturity": years_to_maturity,
            "current_price": current_price,
            "payment_frequency": f"{'Annual' if payment_frequency == 1 else 'Semi-Annual'}",
            "metrics": {
                "ytm_pct": round(ytm_annual, 3),
                "current_yield_pct": round(coupon / current_price * 100, 3) if current_price else 0,
                "macaulay_duration_years": round(mac_dur, 3),
                "modified_duration": round(mod_dur, 3),
                "convexity": round(convexity, 5),
            },
            "price_sensitivity": {
                "price_change_per_1pct_yield_up_pct": round(-mod_dur * 1 + 0.5 * convexity * (1 ** 2), 3),
                "price_change_per_1pct_yield_down_pct": round(mod_dur * 1 + 0.5 * convexity * (1 ** 2), 3),
            },
            "interpretation": f"Macaulay duration of {round(mac_dur, 1)} years means the bond's cash flows are recovered in {round(mac_dur, 1)} years on average." if mac_dur else "",
        }
    except Exception as e:
        logger.error(f"calculate_bond_metrics error: {e}")
        return {"error": str(e)}


async def tool_run_financial_ratio_analysis(symbol: str) -> Dict[str, Any]:
    """Comprehensive financial ratio analysis across 5 categories."""
    try:
        company = await _resolve_company(symbol)
        if not company:
            return {"error": f"Company not found: {symbol}"}

        async with AsyncSessionLocal() as session:
            fq = await session.execute(
                select(Fundamental).where(Fundamental.company_id == company.id)
                .order_by(desc(Fundamental.id)).limit(1)
            )
            fund = fq.scalar_one_or_none()
            pq = await session.execute(
                select(StockQuote).where(StockQuote.company_id == company.id)
                .order_by(desc(StockQuote.timestamp)).limit(1)
            )
            quote = pq.scalar_one_or_none()

        if not fund:
            return {"error": "No fundamental data available"}

        current_price = quote.close if quote else 0
        market_cap = company.market_cap or 0
        revenue = fund.revenue or 0
        profit = fund.profit or 0
        equity = market_cap / (fund.pb or 1) if fund.pb else 0

        ratios = {
            "profitability": {
                "net_margin_pct": fund.net_margin,
                "gross_margin_pct": fund.gross_margin,
                "roe_pct": fund.roe,
                "roce_pct": fund.roce,
                "return_on_assets_pct": round(fund.roe / (1 + (fund.debt_equity or 0)) if fund.roe and fund.debt_equity is not None else 0, 2) if fund.roe else None,
                "eps": fund.eps,
            },
            "liquidity": {
                "current_ratio": fund.current_ratio,
                "quick_ratio_estimate": round((fund.current_ratio or 0) * 0.7, 2) if fund.current_ratio else None,
            },
            "leverage": {
                "debt_to_equity": fund.debt_equity,
                "debt_ratio_estimate": round((fund.debt_equity or 0) / (1 + (fund.debt_equity or 0)), 3) if fund.debt_equity is not None else None,
                "interest_coverage_estimate": round((profit or 0) / ((fund.debt_equity or 0.5) * (revenue or 1) * 0.08 + 1), 2) if revenue else None,
            },
            "efficiency": {
                "asset_turnover_estimate": round(revenue / (equity * (1 + (fund.debt_equity or 0))) if equity > 0 and fund.debt_equity is not None else 0, 3) if revenue and equity > 0 else None,
                "revenue_per_employee_estimate": None,
            },
            "valuation": {
                "pe_ratio": fund.pe,
                "pb_ratio": fund.pb,
                "dividend_yield_pct": fund.dividend_yield,
                "price_to_sales_ratio": round(market_cap / revenue, 2) if revenue > 0 else None,
                "peg_ratio_estimate": round(fund.pe / (fund.roe or 1), 2) if fund.pe and fund.roe else None,
            },
        }

        # Score
        score = 0
        if fund.net_margin and fund.net_margin > 10: score += 1
        if fund.roe and fund.roe > 15: score += 1
        if fund.debt_equity is not None and fund.debt_equity < 0.5: score += 1
        if fund.current_ratio and fund.current_ratio > 1.5: score += 1
        if fund.pe and fund.pe < 25: score += 1
        if fund.dividend_yield and fund.dividend_yield > 1: score += 1

        rating = "Strong Buy" if score >= 6 else "Buy" if score >= 4 else "Hold" if score >= 3 else "Caution"

        return {
            "symbol": symbol,
            "name": company.name,
            "sector": company.sector,
            "current_price": current_price,
            "market_cap_cr": round(market_cap / 1e7, 2),
            "ratios": ratios,
            "health_score": {"score": score, "max": 6, "rating": rating},
            "interpretation": f"Health score {score}/6 — {rating}. {_ratio_interpretation(ratios, fund)}",
        }
    except Exception as e:
        logger.error(f"run_financial_ratio_analysis error: {e}")
        return {"error": str(e)}


def _ratio_interpretation(ratios: dict, fund) -> str:
    parts = []
    p = ratios["profitability"]
    if p.get("net_margin_pct") and p["net_margin_pct"] > 15:
        parts.append("Strong profitability with high margins.")
    elif p.get("net_margin_pct") and p["net_margin_pct"] < 5:
        parts.append("Thin profit margins — cost pressures or pricing weakness.")
    l = ratios["leverage"]
    if l.get("debt_to_equity") is not None and l["debt_to_equity"] < 0.3:
        parts.append("Low financial leverage — conservative balance sheet.")
    elif l.get("debt_to_equity") is not None and l["debt_to_equity"] > 1.5:
        parts.append("High leverage — elevated financial risk.")
    v = ratios["valuation"]
    if v.get("pe_ratio") and v["pe_ratio"] < 15:
        parts.append("Trading at a discount to market (low P/E).")
    elif v.get("pe_ratio") and v["pe_ratio"] > 35:
        parts.append("Premium valuation — high growth expectations priced in.")
    if p.get("roe_pct") and p["roe_pct"] > 20:
        parts.append("Excellent ROE — efficient capital allocation.")
    return " ".join(parts) if parts else "Mixed signals — examine individual ratios."


async def tool_generate_screener(description: str) -> Dict[str, Any]:
    try:
        from openai import AsyncOpenAI

        client = AsyncOpenAI()

        schema = {
            "type": "object",
            "properties": {
                "title": {"type": "string", "description": "Short title for this screener"},
                "filters": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "properties": {
                            "field": {
                                "type": "string",
                                "enum": [
                                    "symbol", "name", "sector", "industry",
                                    "market_cap_cr", "price", "volume",
                                    "pe", "pb", "roe", "roce", "eps",
                                    "debt_equity", "current_ratio",
                                    "gross_margin", "net_margin",
                                    "dividend_yield", "revenue", "profit",
                                ],
                                "description": "The field to filter on",
                            },
                            "op": {
                                "type": "string",
                                "enum": ["eq", "neq", "gte", "lte", "gt", "lt", "contains", "between"],
                                "description": "Comparison operator",
                            },
                            "value": {
                                "description": "Filter value. For 'between' use [min, max]. For 'contains' use a string. For others use a number.",
                            },
                        },
                        "required": ["field", "op", "value"],
                    },
                },
                "logic": {
                    "type": "string",
                    "enum": ["AND", "OR"],
                    "description": "How to combine filters",
                },
                "columns": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "Columns to display in results. Default: symbol, name, sector, price, pe, roe, market_cap_cr",
                },
                "sort_field": {
                    "type": "string",
                    "description": "Field to sort results by",
                },
                "sort_direction": {
                    "type": "string",
                    "enum": ["asc", "desc"],
                    "description": "Sort direction",
                },
                "limit": {
                    "type": "integer",
                    "description": "Maximum results to return (default 20, max 100)",
                },
            },
            "required": ["title", "filters"],
        }

        resp = await client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {
                    "role": "system",
                    "content": (
                        "You generate screener configurations for the Indian stock market. "
                        "Convert natural language descriptions into structured filter JSON. "
                        "Use the correct field names from the available fields list. "
                        "For market cap, use 'market_cap_cr' in crores. "
                        "For percentage values like ROE, gross margin, net margin, use the raw percentage number. "
                        "PE, PB, debt_equity, current_ratio are ratio values. "
                        "Always pick reasonable defaults for columns and sorting if not specified."
                    ),
                },
                {"role": "user", "content": description},
            ],
            response_format={"type": "json_object", "json_schema": {"name": "screener_config", "schema": schema}},
            temperature=0.1,
            max_tokens=1000,
        )

        raw = resp.choices[0].message.content
        if not raw:
            return {"error": "No response from model"}

        config = json.loads(raw)
        config.setdefault("columns", ["symbol", "name", "sector", "price", "pe", "roe", "market_cap_cr"])
        config.setdefault("logic", "AND")
        config.setdefault("sort_field", None)
        config.setdefault("sort_direction", "desc")
        config.setdefault("limit", 20)
        return config
    except Exception as e:
        logger.error(f"generate_screener error: {e}")
        return {"error": str(e)}


TOOL_EXECUTORS = {
    "get_quote": tool_get_quote,
    "get_fundamentals": tool_get_fundamentals,
    "get_historical": tool_get_historical,
    "get_option_chain": tool_get_option_chain,
    "run_risk_analysis": tool_run_risk_analysis,
    "get_company_graph": tool_get_company_graph,
    "search_news": tool_search_news,
    "run_screener": tool_run_screener,
    "get_prediction": tool_get_prediction,
    "get_trading_signals": tool_get_trading_signals,
    "get_sentiment": tool_get_sentiment,
    "compare_stocks": tool_compare_stocks,
    "screen_technicals": tool_screen_technicals,
    "backtest_strategy": tool_backtest_strategy,
    "search_documents": tool_search_documents,
    "get_institutional_activity": tool_get_institutional_activity,
    "get_earnings_calendar": tool_get_earnings_calendar,
    # Excel tools
    "export_analysis_to_excel": tool_export_analysis_to_excel,
    "import_financial_data": tool_import_financial_data,
    # CFA-level tools
    "run_dcf_valuation": tool_run_dcf_valuation,
    "run_comps_analysis": tool_run_comps_analysis,
    "run_dupont_analysis": tool_run_dupont_analysis,
    "calculate_wacc": tool_calculate_wacc,
    "run_portfolio_optimization": tool_run_portfolio_optimization,
    "calculate_bond_metrics": tool_calculate_bond_metrics,
    "run_financial_ratio_analysis": tool_run_financial_ratio_analysis,
    "generate_screener_config": tool_generate_screener,
}
